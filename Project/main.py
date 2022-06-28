import pandas as pd
import xlrd
import openpyxl
from geopy.geocoders import Nominatim
from pprint import pprint

wb = openpyxl.load_workbook('AzatF.xlsx')
eng = wb['Eng']
mag = wb['Mag']
nominaltim = Nominatim(user_agent='user')
# coor = '55.7603, 49.19055'
# location = nominaltim.reverse(coor)
# print(location)
# eng['B3'] = str(location)
# locat = nominaltim.geocode(location).raw
# pprint(locat)
# c = dict(locat)
c = eng.cell(row=1, column=10).value
print(c)
print(mag.max_row)
for i in range(0, mag.max_row - 1):
    print(mag.max_row)
    print(i)
    x = mag.cell(row=2 + i, column=3).value
    print(x)
    y = mag.cell(row=2 + i, column=4).value
    print(y)
    cor = x, y
    location = nominaltim.reverse(cor)
    print(location)
    mag.cell(row=2 + i, column=2).value = str(location)

    # row[6] = 145
    # c = dict(location)





# cor = (eng['C3'].value, eng['D3'].value)
# print(cor)
# location = nominaltim.reverse(cor)
# eng['B3'] = str(location)
# print(location)
wb.save('AzatF.xlsx')



